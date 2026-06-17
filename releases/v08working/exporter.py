import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── цвета ─────────────────────────────────────────────────────────────────────
COLOR_HEADER_BG  = "2F75B6"   # синий заголовок
COLOR_HEADER_FG  = "FFFFFF"   # белый текст
COLOR_ONLINE_BG  = "E8FFE8"   # светло-зелёный
COLOR_ONLINE_FG  = "1A6B1A"   # тёмно-зелёный
COLOR_OFFLINE_BG = "FFE8E8"   # светло-красный
COLOR_OFFLINE_FG = "AA0000"   # тёмно-красный
COLOR_SUMMARY_BG = "F2F2F2"   # серый итог


def _thin_border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def export_to_excel(results: list, vlan: str = "UNKNOWN") -> str:
    """
    Сохраняет результаты скана в Excel.

    Параметры:
        results — список кортежей (ip, hostname, ping, status, mac)
        vlan    — название VLAN для имени файла ('150', 'ALL' и т.д.)

    Возвращает путь к сохранённому файлу.
    """
    os.makedirs("exports", exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename  = f"exports/VLAN{vlan}_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Scan Results"

    # ── заголовки ─────────────────────────────────────────────────────────────
    headers = ["IP", "Hostname", "Ping", "Status", "MAC"]
    ws.append(headers)

    header_font  = Font(bold=True, color=COLOR_HEADER_FG, name="Segoe UI", size=11)
    header_fill  = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, len(headers) + 1):
        cell           = ws.cell(row=1, column=col_idx)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = _thin_border()

    ws.row_dimensions[1].height = 22

    # ── данные ────────────────────────────────────────────────────────────────
    online_fill  = PatternFill("solid", fgColor=COLOR_ONLINE_BG)
    offline_fill = PatternFill("solid", fgColor=COLOR_OFFLINE_BG)
    online_font  = Font(color=COLOR_ONLINE_FG,  name="Segoe UI", size=10)
    offline_font = Font(color=COLOR_OFFLINE_FG, name="Segoe UI", size=10)
    center       = Alignment(horizontal="center")

    online_count = offline_count = 0

    for row_data in results:
        # поддержка как 4-элементных (старых) так и 5-элементных кортежей
        if len(row_data) == 4:
            ip, hostname, ping, status = row_data
            mac = "N/A"
        else:
            ip, hostname, ping, status, mac = row_data

        ws.append([ip, hostname, ping, status, mac])
        row_idx   = ws.max_row
        is_online = str(status).upper() == "ONLINE"

        if is_online:
            online_count += 1
        else:
            offline_count += 1

        fill = online_fill  if is_online else offline_fill
        font = online_font  if is_online else offline_font

        for col_idx in range(1, len(headers) + 1):
            cell           = ws.cell(row=row_idx, column=col_idx)
            cell.fill      = fill
            cell.font      = font
            cell.border    = _thin_border()
            if col_idx in (3, 4):      # Ping и Status — по центру
                cell.alignment = center

    # ── итоговая строка ───────────────────────────────────────────────────────
    summary_row = ws.max_row + 1
    ws.cell(row=summary_row, column=1, value="TOTAL")
    ws.cell(row=summary_row, column=2, value=f"{len(results)} hosts")
    ws.cell(row=summary_row, column=3, value=f"🟢 {online_count}")
    ws.cell(row=summary_row, column=4, value=f"🔴 {offline_count}")
    ws.cell(row=summary_row, column=5, value="")

    summary_fill = PatternFill("solid", fgColor=COLOR_SUMMARY_BG)
    summary_font = Font(bold=True, name="Segoe UI", size=10)

    for col_idx in range(1, len(headers) + 1):
        cell           = ws.cell(row=summary_row, column=col_idx)
        cell.fill      = summary_fill
        cell.font      = summary_font
        cell.border    = _thin_border()
        cell.alignment = Alignment(horizontal="center")

    # ── авто-ширина колонок ───────────────────────────────────────────────────
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    # ── заморозить первую строку ──────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── авто-фильтр ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:E{ws.max_row - 1}"

    wb.save(filename)
    return filename